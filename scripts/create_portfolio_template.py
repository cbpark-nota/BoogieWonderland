"""
portfolio_template.xlsx 와 portfolio.xlsx (샘플 데이터 포함) 생성
─────────────────────────────────────────────────────────────────
사용법:
  python scripts/create_portfolio_template.py

컬럼:
  ticker      | 티커 (US: AAPL, KR: 005930.KS)
  name        | 종목명 (선택, 비워두면 자동 조회)
  market      | 시장 (US / KR)
  entry_price | 진입가
  shares      | 주수
  entry_date  | 진입일 (YYYY-MM-DD)
  stop_loss   | 스톱로스 가격 (선택)
  target_price| 목표가 (선택)
  memo        | 메모 (선택)
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("ticker",       "티커",           18),
    ("name",         "종목명",          20),
    ("market",       "시장(US/KR)",    12),
    ("entry_price",  "진입가",          14),
    ("shares",       "주수",           10),
    ("entry_date",   "진입일",          14),
    ("stop_loss",    "스톱로스",        14),
    ("target_price", "목표가",          14),
    ("memo",         "메모",           30),
]

SAMPLE_ROWS = [
    ["AAPL",       "Apple Inc.",   "US", 200.00, 10, "2026-01-15", 180.00, 250.00, ""],
    ["NVDA",       "NVIDIA Corp.", "US", 130.00,  5, "2026-02-01", 110.00, 180.00, "AI 반도체"],
    ["005930.KS",  "삼성전자",      "KR", 80000,  5, "2026-01-20", 72000, 100000, ""],
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")


def create_workbook(rows: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # 헤더
    for col_idx, (_, label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, row_data in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left")

    # 틀 고정 (헤더 아래)
    ws.freeze_panes = "A2"

    return wb


def main():
    root = Path(__file__).parent
    template_path = root / "portfolio_template.xlsx"
    portfolio_path = root / "portfolio.xlsx"

    # 템플릿 (샘플 데이터 포함)
    wb = create_workbook(SAMPLE_ROWS)
    wb.save(template_path)
    print(f"템플릿 생성: {template_path}")

    # portfolio.xlsx — 리포지토리에 포함할 초기 파일 (샘플 데이터 포함)
    if not portfolio_path.exists():
        wb2 = create_workbook(SAMPLE_ROWS)
        wb2.save(portfolio_path)
        print(f"포트폴리오 초기 파일 생성: {portfolio_path}")
    else:
        print(f"포트폴리오 파일 이미 존재 (덮어쓰지 않음): {portfolio_path}")


if __name__ == "__main__":
    main()
